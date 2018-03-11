from django.shortcuts import render,redirect
from rest_framework import viewsets
from rest_framework.response import Response
from django.http import HttpResponse,HttpRequest
from api import models
from api.serializers import userSerializer,ITSerializer,userlistSerializer,CSSerializer,subSerializer,css1Serializer,its1Serializer,AvaiSerializer
from api.serializers import its3Serializer,css3Serializer,css4Serializer,its4Serializer
import csv,json,os
from tabula import read_pdf,convert_into
from pprint import pprint
import requests
from openpyxl import Workbook
from django.core.management import call_command
from django.db import models as mymodel
from rest_framework import generics
from django.views.decorators.csrf import csrf_exempt
import os
from django.core.files.storage import default_storage
from django.core.files.base import ContentFile
from django.conf import settings
from django.db import IntegrityError
from django.core.exceptions import ObjectDoesNotExist
from django.views.static import serve
from wsgiref.util import FileWrapper



class userViewSet(viewsets.ModelViewSet):
    queryset = models.userlist.objects.all()
    serializer_class = userSerializer

class ITViewSet(viewsets.ModelViewSet):
    queryset = models.ITS2.objects.all()
    serializer_class = ITSerializer

class CSViewSet(viewsets.ModelViewSet):
    queryset = models.CSS2.objects.all()
    serializer_class = CSSerializer

class it1ViewSet(viewsets.ModelViewSet):
    queryset=models.ITS1.objects.all()
    serializer_class=its1Serializer

class cs1ViewSet(viewsets.ModelViewSet):
    queryset=models.CSS1.objects.all()
    serializer_class=css1Serializer

class AvaiViewSet(viewsets.ModelViewSet):
    queryset=models.available.objects.all()
    serializer_class=AvaiSerializer

class AvaiSemViewSet(generics.ListAPIView):
    serializer_class=AvaiSerializer
    def get_queryset(self):
        semester=self.kwargs['semester']
        return models.available.objects.filter(semester=semester)


class markList(generics.ListAPIView):
    def get_queryset(self):
        department=self.kwargs['department']
        semester=self.kwargs['semester']
        if(department=='CS'):
            if(semester=='1'):
                return models.CSS1.objects.all()
            if(semester=='2'):
                return models.CSS2.objects.all()
            if(semester=='3'):
                return models.CSS3.objects.all()
            if(semester=='4'):
                return models.CSS4.objects.all()
        if(department=='IT'):
            if(semester=='1'):
                return models.ITS1.objects.all()
            if(semester=='2'):
                return models.ITS2.objects.all()
            if(semester=='3'):
                return models.ITS3.objects.all()
            if(semester=='4'):
                return models.ITS4.objects.all()

    def get_serializer_class(self):
        department=self.kwargs['department']
        semester=self.kwargs['semester']
        if(department=='CS'):
            if(semester=='1'):
                return css1Serializer
            if(semester=='2'):
                return CSSerializer
            if(semester=='3'):
                return css3Serializer
            if(semester=='4'):
                return css4Serializer
        if(department=='IT'):
            if(semester=='1'):
                return its1Serializer
            if(semester=='2'):
                return ITSerializer
            if(semester=='3'):
                return its3Serializer
            if(semester=='4'):
                return its4Serializer

class markListBatch(generics.ListAPIView):
    def get_queryset(self):
        department=self.kwargs['department']
        semester=self.kwargs['semester']
        batch=self.kwargs['batch']
        if(department=='CS'):
            if(semester=='1'):
                return models.CSS1.objects.filter(batch=batch)
            if(semester=='2'):
                return models.CSS2.objects.filter(batch=batch)
            if(semester=='3'):
                return models.CSS3.objects.filter(batch=batch)
            if(semester=='4'):
                return models.CSS4.objects.filter(batch=batch)
        if(department=='IT'):
            if(semester=='1'):
                return models.ITS1.objects.filter(batch=batch)
            if(semester=='2'):
                return models.ITS2.objects.filter(batch=batch)
            if(semester=='3'):
                return models.ITS3.objects.filter(batch=batch)
            if(semester=='4'):
                return models.ITS4.objects.filter(batch=batch)

    def get_serializer_class(self):
        department=self.kwargs['department']
        semester=self.kwargs['semester']
        if(department=='CS'):
            if(semester=='1'):
                return css1Serializer
            if(semester=='2'):
                return CSSerializer
            if(semester=='3'):
                return css3Serializer
            if(semester=='4'):
                return css4Serializer
        if(department=='IT'):
            if(semester=='1'):
                return its1Serializer
            if(semester=='2'):
                return ITSerializer
            if(semester=='3'):
                return its3Serializer
            if(semester=='4'):
                return its4Serializer


class markListRoll(generics.ListCreateAPIView):
    def get_queryset(self):
        department=self.kwargs['department']
        semester=self.kwargs['semester']
        rollno=self.kwargs['rollno']
        if(department=='CS'):
            if(semester=='1'):
                return models.CSS1.objects.filter(rollno=rollno)
            if(semester=='2'):
                return models.CSS2.objects.filter(rollno=rollno)
            if(semester=='3'):
                return models.CSS3.objects.filter(rollno=rollno)
            if(semester=='4'):
                return models.CSS4.objects.filter(rollno=rollno)
        if(department=='IT'):
            if(semester=='1'):
                return models.ITS1.objects.filter(rollno=rollno)
            if(semester=='2'):
                return models.ITS2.objects.filter(rollno=rollno)
            if(semester=='3'):
                return models.ITS3.objects.filter(rollno=rollno)
            if(semester=='4'):
                return models.ITS4.objects.filter(rollno=rollno)

    def get_serializer_class(self):
        department=self.kwargs['department']
        semester=self.kwargs['semester']
        if(department=='CS'):
            if(semester=='1'):
                return css1Serializer
            if(semester=='2'):
                return CSSerializer
            if(semester=='3'):
                return css3Serializer
            if(semester=='4'):
                return css4Serializer
        if(department=='IT'):
            if(semester=='1'):
                return its1Serializer
            if(semester=='2'):
                return ITSerializer
            if(semester=='3'):
                return its3Serializer
            if(semester=='4'):
                return its4Serializer



class userlistViewSet(viewsets.ModelViewSet):
    queryset = models.userlist.objects.all()
    serializer_class = userlistSerializer

class PurchaseList(generics.ListAPIView):
    serializer_class = subSerializer

    def get_queryset(self):
        """
        This view should return a list of all the purchases for
        the user as determined by the username portion of the URL.
        """
        department = self.kwargs['department']
        semester = self.kwargs['semester']
        return models.subjectList.objects.filter(department=department).filter(semester=semester)


class subjectViewSet(viewsets.ModelViewSet):
    queryset= models.subjectList.objects.all()
    serializer_class = subSerializer

@csrf_exempt
def datasave(request,semester,batch):
    try:
        oo=models.available.objects.create(semester=semester,batch=batch)
        oo.save()
    except IntegrityError as e:
        return HttpResponse(status=400)
    name='tmp/sem'+semester+'batch'+batch+'.pdf'
    if request.method == "POST":
        aa=request.FILES['pdf']
        path = default_storage.save(name, ContentFile(aa.read()))
        tmp_file = os.path.join(settings.MEDIA_ROOT, path)
    data={}
    code={}
    convert_into('tmp/sem'+semester+'batch'+batch+'.pdf',"output.csv", output_format="csv",lattice="True",pages="all",header=None)
    with open("output.csv","r") as f,open('data.json', 'w+') as outfile:
        reader = csv.reader(f)
        for row in reader:
            if (len(row[0])==10 or len(row[0])==11):
                row[1]=row[1].replace("\n","")
                data.setdefault("data",[]).append({"rollno": row[0],"mark": row[1]})
        json.dump(data,outfile, indent=4, sort_keys=True)
    with open("output.csv","r") as f,open('code.json', 'w+') as outfile:
        reader = csv.reader(f)
        for row in reader:
            if len(row[0])==5:
                code.setdefault("code",[]).append({"code": row[0],"subject": row[1]})
        json.dump(code,outfile, indent=4, sort_keys=True)
    return redirect('http://127.0.0.1:8000/apii/extractall/'+semester+'/'+batch+'/')

def extractall(request,semester,batch):
    json_data=open('./data.json')
    data1 = json.load(json_data)
    if(semester=='2'):
        print('inside')
        i=0
        for a in data1['data']:
            roll=data1['data'][i]['rollno']
            w=data1['data'][i]['mark']
            w=[x.strip() for x in w.split(',')]
            k=0
            s=list(roll)
            if (len(s)==10):
                b=s[3]+s[4]
            elif (len(s)==11):
                b=s[4]+s[5]
            if(b!=batch):
                i+=1
                continue
            batch=b
            if((s[5]=='I' and s[6]=='T') or (s[6]=='I' and s[7]=='T')):
                for y in w:
                    b=list(w[k])
                    #PH100=MA102=BE110=BE102=PH110=EC110=EC100=CS120=CS100='absent'
                    if "PH100" in w[k]:
                        if len(b)==8:
                            grade=b[6]
                        elif len(b)==9:
                            grade=b[6]+b[7]
                        else:
                            grade='absent'
                        PH100=grade
                    if "MA102" in w[k]:
                        if len(b)==8:
                            grade=b[6]
                        elif len(b)==9:
                            grade=b[6]+b[7]
                        else:
                            grade='absent'
                        MA102=grade
                    if "BE110" in w[k]:
                        if len(b)==8:
                            grade=b[6]
                        elif len(b)==9:
                            grade=b[6]+b[7]
                        else:
                            grade='absent'
                        BE110=grade
                    if "BE102" in w[k]:
                        if len(b)==8:
                            grade=b[6]
                        elif len(b)==9:
                            grade=b[6]+b[7]
                        else:
                            grade='absent'
                        BE102=grade
                    if "PH110" in w[k]:
                        if len(b)==8:
                            grade=b[6]
                        elif len(b)==9:
                            grade=b[6]+b[7]
                        else:
                            grade='absent'
                        PH110=grade
                    if "EC100" in w[k]:
                        if len(b)==8:
                            grade=b[6]
                        elif len(b)==9:
                            grade=b[6]+b[7]
                        else:
                            grade='absent'
                        EC100=grade
                    if "EC110" in w[k]:
                        if len(b)==8:
                            grade=b[6]
                        elif len(b)==9:
                            grade=b[6]+b[7]
                        else:
                            grade='absent'
                        EC110=grade
                    if "CS120" in w[k]:
                        if len(b)==8:
                            grade=b[6]
                        elif len(b)==9:
                            grade=b[6]+b[7]
                        else:
                            grade='absent'
                        CS120=grade
                    if "CS100" in w[k]:
                        if len(b)==8:
                            grade=b[6]
                        elif len(b)==9:
                            grade=b[6]+b[7]
                        else:
                            grade='absent'
                        CS100=grade
                    k=k+1
                Subs = models.ITS2.objects.create(rollno=roll,PH100=PH100,MA102=MA102,BE110=BE110,BE102=BE102,PH110=PH110,EC100=EC100,EC110=EC110,CS120=CS120,CS100=CS100,batch=batch)
                Subs.save()
            if((s[5]=='C' and s[6]=='S') or (s[6]=='C' and s[7]=='S')):
                for y in w:
                    b=list(w[k])
                    #MA102=PH100=BE110=BE102=PH110=EE100=EE110=CS120=CS100='absent'
                    if "MA102" in w[k]:
                        if len(b)==8:
                            grade=b[6]
                        elif len(b)==9:
                            grade=b[6]+b[7]
                        else:
                            grade='absent'
                        PH100=grade
                    if "PH100" in w[k]:
                        if len(b)==8:
                            grade=b[6]
                        elif len(b)==9:
                            grade=b[6]+b[7]
                        else:
                            grade='absent'
                        MA102=grade
                    if "BE110" in w[k]:
                        if len(b)==8:
                            grade=b[6]
                        elif len(b)==9:
                            grade=b[6]+b[7]
                        else:
                            grade='absent'
                        BE110=grade
                    if "BE102" in w[k]:
                        if len(b)==8:
                            grade=b[6]
                        elif len(b)==9:
                            grade=b[6]+b[7]
                        else:
                            grade='absent'
                        BE102=grade
                    if "PH110" in w[k]:
                        if len(b)==8:
                            grade=b[6]
                        elif len(b)==9:
                            grade=b[6]+b[7]
                        else:
                            grade='absent'
                        PH110=grade
                    if "EE100" in w[k]:
                        if len(b)==8:
                            grade=b[6]
                        elif len(b)==9:
                            grade=b[6]+b[7]
                        else:
                            grade='absent'
                        EC100=grade
                    if "EE110" in w[k]:
                        if len(b)==8:
                            grade=b[6]
                        elif len(b)==9:
                            grade=b[6]+b[7]
                        else:
                            grade='absent'
                        EC110=grade
                    if "CS120" in w[k]:
                        if len(b)==8:
                            grade=b[6]
                        elif len(b)==9:
                            grade=b[6]+b[7]
                        else:
                            grade='absent'
                        CS120=grade
                    if "CS100" in w[k]:
                        if len(b)==8:
                            grade=b[6]
                        elif len(b)==9:
                            grade=b[6]+b[7]
                        else:
                            grade='absent'
                        CS100=grade
                    k=k+1
                Subs = models.CSS2.objects.create(rollno=roll,PH100=PH100,MA102=MA102,BE110=BE110,BE102=BE102,PH110=PH110,EE100=EC100,EE110=EC110,CS120=CS120,CS100=CS100,batch=batch)
                Subs.save()
            i=i+1
    elif(semester=='1'):
        i=0
        for a in data1['data']:
            roll=data1['data'][i]['rollno']
            w=data1['data'][i]['mark']
            w=[x.strip() for x in w.split(',')]
            k=0
            s=list(roll)
            if (len(s)==10):
                b=s[3]+s[4]
            elif (len(s)==11):
                b=s[4]+s[5]
            if(b!=batch):
                i+=1
                continue
            batch=b
            if((s[5]=='I' and s[6]=='T') or (s[6]=='I' and s[7]=='T')):
                for y in w:
                    b=list(w[k])
                    if "MA101" in w[k]:
                        if len(b)==8:
                            grade=b[6]
                        elif len(b)==9:
                            grade=b[6]+b[7]
                        else:
                            grade='absent'
                        MA101=grade
                    if "CY100" in w[k]:
                        if len(b)==8:
                            grade=b[6]
                        elif len(b)==9:
                            grade=b[6]+b[7]
                        else:
                            grade='absent'
                        CY100=grade
                    if "BE100" in w[k]:
                        if len(b)==8:
                            grade=b[6]
                        elif len(b)==9:
                            grade=b[6]+b[7]
                        else:
                            grade='absent'
                        BE100=grade
                    if "BE102" in w[k]:
                        if len(b)==8:
                            grade=b[6]
                        elif len(b)==9:
                            grade=b[6]+b[7]
                        else:
                            grade='absent'
                        BE102=grade
                    if "BE10105" in w[k]:
                        if len(b)==10:
                            grade=b[8]
                        elif len(b)==11:
                            grade=b[8]+b[9]
                        else:
                            grade='absent'
                        BE10105=grade
                    if "EE100" in w[k]:
                        if len(b)==8:
                            grade=b[6]
                        elif len(b)==9:
                            grade=b[6]+b[7]
                        else:
                            grade='absent'
                        EE100=grade
                    if "BE110" in w[k]:
                        if len(b)==8:
                            grade=b[6]
                        elif len(b)==9:
                            grade=b[6]+b[7]
                        else:
                            grade='absent'
                        BE110=grade
                    if "EC100" in w[k]:
                        if len(b)==8:
                            grade=b[6]
                        elif len(b)==9:
                            grade=b[6]+b[7]
                        else:
                            grade='absent'
                        EC100=grade
                    if "BE103" in w[k]:
                        if len(b)==8:
                            grade=b[6]
                        elif len(b)==9:
                            grade=b[6]+b[7]
                        else:
                            grade='absent'
                        BE103=grade
                    if "CY110" in w[k]:
                        if len(b)==8:
                            grade=b[6]
                        elif len(b)==9:
                            grade=b[6]+b[7]
                        else:
                            grade='absent'
                        CY110=grade
                    if "CS110" in w[k]:
                        if len(b)==8:
                            grade=b[6]
                        elif len(b)==9:
                            grade=b[6]+b[7]
                        else:
                            grade='absent'
                        CS110=grade
                    if "EE110" in w[k]:
                        if len(b)==8:
                            grade=b[6]
                        elif len(b)==9:
                            grade=b[6]+b[7]
                        else:
                            grade='absent'
                        EE110=grade
                    k=k+1
                Subs = models.ITS1.objects.create(rollno=roll,MA101=MA101,CY100=CY100,BE100=BE100,BE10105=BE10105,EE100=EE100,BE103=BE103,CY110=CY110,CS110=CS110,EE110=EE110,batch=batch)
                Subs.save()
            if((s[5]=='C' and s[6]=='S') or (s[6]=='C' and s[7]=='S')):
                for y in w:
                    b=list(w[k])
                    #MA102=PH100=BE110=BE102=PH110=EE100=EE110=CS120=CS100='absent'
                    if "MA101" in w[k]:
                        if len(b)==8:
                            grade=b[6]
                        elif len(b)==9:
                            grade=b[6]+b[7]
                        else:
                            grade='absent'
                        MA101=grade
                    if "CY100" in w[k]:
                        if len(b)==8:
                            grade=b[6]
                        elif len(b)==9:
                            grade=b[6]+b[7]
                        else:
                            grade='absent'
                        CY100=grade
                    if "BE100" in w[k]:
                        if len(b)==8:
                            grade=b[6]
                        elif len(b)==9:
                            grade=b[6]+b[7]
                        else:
                            grade='absent'
                        BE100=grade
                    if "BE10105" in w[k]:
                        if len(b)==10:
                            grade=b[8]
                        elif len(b)==11:
                            grade=b[8]+b[9]
                        else:
                            grade='absent'
                        BE10105=grade
                    if "EC100" in w[k]:
                        if len(b)==8:
                            grade=b[6]
                        elif len(b)==9:
                            grade=b[6]+b[7]
                        else:
                            grade='absent'
                        EC100=grade
                    if "BE110" in w[k]:
                        if len(b)==8:
                            grade=b[6]
                        elif len(b)==9:
                            grade=b[6]+b[7]
                        else:
                            grade='absent'
                        BE110=grade
                    if "BE103" in w[k]:
                        if len(b)==8:
                            grade=b[6]
                        elif len(b)==9:
                            grade=b[6]+b[7]
                        else:
                            grade='absent'
                        BE103=grade
                    if "CY110" in w[k]:
                        if len(b)==8:
                            grade=b[6]
                        elif len(b)==9:
                            grade=b[6]+b[7]
                        else:
                            grade='absent'
                        CY110=grade
                    if "CS110" in w[k]:
                        if len(b)==8:
                            grade=b[6]
                        elif len(b)==9:
                            grade=b[6]+b[7]
                        else:
                            grade='absent'
                        CS110=grade
                    if "EC110" in w[k]:
                        if len(b)==8:
                            grade=b[6]
                        elif len(b)==9:
                            grade=b[6]+b[7]
                        else:
                            grade='absent'
                        EC110=grade
                    k=k+1
                Subs = models.CSS1.objects.create(rollno=roll,MA101=MA101,CY100=CY100,BE100=BE100,BE10105=BE10105,EC100=EC100,BE103=BE103,CY110=CY110,CS110=CS110,EC110=EC110,batch=batch)
                Subs.save()
            i=i+1 
    elif(semester=='3'):
        i=0
        for a in data1['data']:
            roll=data1['data'][i]['rollno']
            w=data1['data'][i]['mark']
            w=[x.strip() for x in w.split(',')]
            k=0
            s=list(roll)
            if (len(s)==10):
                b=s[3]+s[4]
            elif (len(s)==11):
                b=s[4]+s[5]
            if(b!=batch):
                i+=1
                continue
            batch=b
            if((s[5]=='I' and s[6]=='T') or (s[6]=='I' and s[7]=='T')):
                for y in w:
                    b=list(w[k])
                    if "MA201" in w[k]:
                        if len(b)==8:
                            grade=b[6]
                        elif len(b)==9:
                            grade=b[6]+b[7]
                        else:
                            grade='absent'
                        MA201=grade
                    if "CS201" in w[k]:
                        if len(b)==8:
                            grade=b[6]
                        elif len(b)==9:
                            grade=b[6]+b[7]
                        else:
                            grade='absent'
                        CS201=grade
                    if "IT201" in w[k]:
                        if len(b)==8:
                            grade=b[6]
                        elif len(b)==9:
                            grade=b[6]+b[7]
                        else:
                            grade='absent'
                        IT201=grade
                    if "CS205" in w[k]:
                        if len(b)==8:
                            grade=b[6]
                        elif len(b)==9:
                            grade=b[6]+b[7]
                        else:
                            grade='absent'
                        CS205=grade
                    if "IT203" in w[k]:
                        if len(b)==8:
                            grade=b[6]
                        elif len(b)==9:
                            grade=b[6]+b[7]
                        else:
                            grade='absent'
                        IT203=grade
                    if "BE110" in w[k]:
                        if len(b)==8:
                            grade=b[6]
                        elif len(b)==9:
                            grade=b[6]+b[7]
                        else:
                            grade='absent'
                        BE110=grade
                    if "HS200" in w[k]:
                        if len(b)==8:
                            grade=b[6]
                        elif len(b)==9:
                            grade=b[6]+b[7]
                        else:
                            grade='absent'
                        HS200=grade
                    if "CS231" in w[k]:
                        if len(b)==8:
                            grade=b[6]
                        elif len(b)==9:
                            grade=b[6]+b[7]
                        else:
                            grade='absent'
                        CS231=grade
                    if "IT231" in w[k]:
                        if len(b)==8:
                            grade=b[6]
                        elif len(b)==9:
                            grade=b[6]+b[7]
                        else:
                            grade='absent'
                        IT231=grade
                    k=k+1
                Subs = models.ITS3.objects.create(rollno=roll,MA201=MA201,CS201=CS201,IT201=IT201,CS205=CS205,IT203=IT203,HS200=HS200,CS231=CS231,IT231=IT231,batch=batch)
                Subs.save()
            if((s[5]=='C' and s[6]=='S') or (s[6]=='C' and s[7]=='S')):
                for y in w:
                    b=list(w[k])
                    #MA102=PH100=BE110=BE102=PH110=EE100=EE110=CS120=CS100='absent'
                    if "MA201" in w[k]:
                        if len(b)==8:
                            grade=b[6]
                        elif len(b)==9:
                            grade=b[6]+b[7]
                        else:
                            grade='absent'
                        MA201=grade
                    if "CS201" in w[k]:
                        if len(b)==8:
                            grade=b[6]
                        elif len(b)==9:
                            grade=b[6]+b[7]
                        else:
                            grade='absent'
                        CS201=grade
                    if "CS203" in w[k]:
                        if len(b)==8:
                            grade=b[6]
                        elif len(b)==9:
                            grade=b[6]+b[7]
                        else:
                            grade='absent'
                        CS203=grade
                    if "CS205" in w[k]:
                        if len(b)==8:
                            grade=b[6]
                        elif len(b)==9:
                            grade=b[6]+b[7]
                        else:
                            grade='absent'
                        CS205=grade
                    if "CS207" in w[k]:
                        if len(b)==8:
                            grade=b[6]
                        elif len(b)==9:
                            grade=b[6]+b[7]
                        else:
                            grade='absent'
                        CS207=grade
                    if "HS200" in w[k]:
                        if len(b)==8:
                            grade=b[6]
                        elif len(b)==9:
                            grade=b[6]+b[7]
                        else:
                            grade='absent'
                        HS200=grade
                    if "CS231" in w[k]:
                        if len(b)==8:
                            grade=b[6]
                        elif len(b)==9:
                            grade=b[6]+b[7]
                        else:
                            grade='absent'
                        CS231=grade
                    if "CS233" in w[k]:
                        if len(b)==8:
                            grade=b[6]
                        elif len(b)==9:
                            grade=b[6]+b[7]
                        else:
                            grade='absent'
                        CS233=grade
                    if "EC110" in w[k]:
                        if len(b)==8:
                            grade=b[6]
                        elif len(b)==9:
                            grade=b[6]+b[7]
                        else:
                            grade='absent'
                        EC110=grade
                    k=k+1
                Subs = models.CSS3.objects.create(rollno=roll,MA201=MA201,CS201=CS201,CS203=CS203,CS205=CS205,CS207=CS207,HS200=HS200,CS231=CS231,CS233=CS233,batch=batch)
                Subs.save()
            i=i+1
    elif(semester=='4'):
        i=0
        for a in data1['data']:
            roll=data1['data'][i]['rollno']
            w=data1['data'][i]['mark']
            w=[x.strip() for x in w.split(',')]
            k=0
            s=list(roll)
            if (len(s)==10):
                b=s[3]+s[4]
            elif (len(s)==11):
                b=s[4]+s[5]
            if(b!=batch):
                i+=1
                continue
            batch=b
            if((s[5]=='I' and s[6]=='T') or (s[6]=='I' and s[7]=='T')):
                for y in w:
                    b=list(w[k])
                    if "MA202" in w[k]:
                        if len(b)==8:
                            grade=b[6]
                        elif len(b)==9:
                            grade=b[6]+b[7]
                        else:
                            grade='absent'
                        MA202=grade
                    if "CS202" in w[k]:
                        if len(b)==8:
                            grade=b[6]
                        elif len(b)==9:
                            grade=b[6]+b[7]
                        else:
                            grade='absent'
                        CS202=grade
                    if "IT202" in w[k]:
                        if len(b)==8:
                            grade=b[6]
                        elif len(b)==9:
                            grade=b[6]+b[7]
                        else:
                            grade='absent'
                        IT202=grade
                    if "IT204" in w[k]:
                        if len(b)==8:
                            grade=b[6]
                        elif len(b)==9:
                            grade=b[6]+b[7]
                        else:
                            grade='absent'
                        IT204=grade
                    if "CS208" in w[k]:
                        if len(b)==8:
                            grade=b[6]
                        elif len(b)==9:
                            grade=b[6]+b[7]
                        else:
                            grade='absent'
                        CS208=grade
                    if "HS210" in w[k]:
                        if len(b)==8:
                            grade=b[6]
                        elif len(b)==9:
                            grade=b[6]+b[7]
                        else:
                            grade='absent'
                        HS210=grade
                    if "IT232" in w[k]:
                        if len(b)==8:
                            grade=b[6]
                        elif len(b)==9:
                            grade=b[6]+b[7]
                        else:
                            grade='absent'
                        IT232=grade
                    if "IT234" in w[k]:
                        if len(b)==8:
                            grade=b[6]
                        elif len(b)==9:
                            grade=b[6]+b[7]
                        else:
                            grade='absent'
                        IT234=grade
                    k=k+1
                Subs = models.ITS4.objects.create(rollno=roll,MA202=MA202,CS202=CS202,IT202=IT202,IT204=IT204,CS208=CS208,HS210=HS210,IT232=IT232,IT234=IT234,batch=batch)
                Subs.save()
            if((s[5]=='C' and s[6]=='S') or (s[6]=='C' and s[7]=='S')):
                for y in w:
                    b=list(w[k])
                    #MA102=PH100=BE110=BE102=PH110=EE100=EE110=CS120=CS100='absent'
                    if "MA202" in w[k]:
                        if len(b)==8:
                            grade=b[6]
                        elif len(b)==9:
                            grade=b[6]+b[7]
                        else:
                            grade='absent'
                        MA202=grade
                    if "CS202" in w[k]:
                        if len(b)==8:
                            grade=b[6]
                        elif len(b)==9:
                            grade=b[6]+b[7]
                        else:
                            grade='absent'
                        CS202=grade
                    if "CS204" in w[k]:
                        if len(b)==8:
                            grade=b[6]
                        elif len(b)==9:
                            grade=b[6]+b[7]
                        else:
                            grade='absent'
                        CS204=grade
                    if "CS206" in w[k]:
                        if len(b)==8:
                            grade=b[6]
                        elif len(b)==9:
                            grade=b[6]+b[7]
                        else:
                            grade='absent'
                        CS206=grade
                    if "CS208" in w[k]:
                        if len(b)==8:
                            grade=b[6]
                        elif len(b)==9:
                            grade=b[6]+b[7]
                        else:
                            grade='absent'
                        CS208=grade
                    if "HS210" in w[k]:
                        if len(b)==8:
                            grade=b[6]
                        elif len(b)==9:
                            grade=b[6]+b[7]
                        else:
                            grade='absent'
                        HS210=grade
                    if "CS232" in w[k]:
                        if len(b)==8:
                            grade=b[6]
                        elif len(b)==9:
                            grade=b[6]+b[7]
                        else:
                            grade='absent'
                        CS232=grade
                    if "CS234" in w[k]:
                        if len(b)==8:
                            grade=b[6]
                        elif len(b)==9:
                            grade=b[6]+b[7]
                        else:
                            grade='absent'
                        CS234=grade
                    k=k+1
                Subs = models.CSS4.objects.create(rollno=roll,MA202=MA202,CS202=CS202,CS204=CS204,CS206=CS206,CS208=CS208,HS210=HS210,CS232=CS232,CS234=CS234,batch=batch)
                Subs.save()
            i=i+1     
    return HttpResponse('success')

def excelCreation(request,semester,batch,department):
    wb = Workbook()
    ws = wb.active
    ws['A1']='Rollno'
    aab='http://127.0.0.1:8000/apii/subs/'+department+'/'+semester+'/'
    r = requests.get(aab)
    for line in r.iter_lines():
        if line:
            co=json.loads(line)
    codes=co
    print(codes)
    aaa='http://127.0.0.1:8000/apii/marksBatch/'+department+'/'+semester+'/'+batch+'/'
    r = requests.get(aaa)
    for line in r.iter_lines():
        if line:
            mrks=json.loads(line)
    marks=mrks
    avai=[]
    Credit=[]
    for i in range(15):
       avai.append(0)
    for i in range(15):
       Credit.append(0)
    x=2
    i=0
    for y in range(0,len(codes)):
        try:
            c=codes[y]['code']
            check=marks[y][c]
            avai[i]=codes[y]['code']
            Credit[i]=codes[y]['credit']
            ws.cell(row=1,column=x).value=codes[y]['code']
            x+=1
            i+=1
        except KeyError as e:
            print("key error")
    ws.cell(row=1,column=x).value="gpa"
    ws.cell(row=1,column=x+1).value="Result"
    ws.cell(row=1,column=x+2).value="No_Arrears"
    ws.cell(row=1,column=x+3).value="ArrearSubjects"
    SemCredit=0
    for i in range(15):
       SemCredit=SemCredit+(Credit[i]*10)
    i=2
    length=len(avai)
    CreditArray={'O':10,"A+":9,"A":8,"B+":7,"B":6,"C":5,"P":4,"F":0,"FA":0,'FE':0,"FS":0,"absent":0}
    for y in range(0,len(marks)):
        g=2
        totalGpa=0
        status="pass"
        NoArrears=0
        ArrearSubjects=''
        for z in range(0,length):
            code=avai[z]
            ws.cell(row=i,column=1).value=marks[y]['rollno']
            try:
                ws.cell(row=i,column=g).value=marks[y][code]
                grade=marks[y][code]
                totalGpa=totalGpa+(Credit[z]*CreditArray[grade])
                g+=1
                if(CreditArray[grade]==0):
                    status="fail"
                    ArrearSubjects=ArrearSubjects+code+','
                    NoArrears+=1
            except KeyError as e:
                print("error")
        gpa=totalGpa/SemCredit
        gpa=round(gpa*10,2)
        if(status=="fail"):
            gpa=0
        ws.cell(row=i,column=g).value=gpa
        ws.cell(row=i,column=g+1).value=status
        ws.cell(row=i,column=g+2).value=NoArrears
        ws.cell(row=i,column=g+3).value=ArrearSubjects
        i+=1
    wb.save('alan.xlsx')
    filename = 'alan.xlsx'
    wrapper = FileWrapper(open(filename,'rb'))
    response = HttpResponse(wrapper, content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = 'attachment; filename=%s' % os.path.basename(filename)
    response['Content-Length'] = os.path.getsize(filename)
    return response

def updatemark(request,semester,department,rollno,grade,code):
        try:
            if(department=='CS'):
                if(semester=='1'):
                    f =models.CSS1.objects.get(rollno=rollno)
                    setattr(f,code,grade)
                    f.save()
                if(semester=='2'):
                    f =models.CSS2.objects.get(rollno=rollno)
                    setattr(f,code,grade)
                    f.save()
                if(semester=='3'):
                    f =models.CSS3.objects.get(rollno=rollno)
                    setattr(f,code,grade)
                    f.save()
                if(semester=='4'):
                    f =models.CSS4.objects.get(rollno=rollno)
                    setattr(f,code,grade)
                    f.save()
            elif(department=='IT'):
                if(semester=='1'):
                    f =models.ITS1.objects.get(rollno=rollno)
                    setattr(f,code,grade)
                    f.save()
                if(semester=='2'):
                    f =models.ITS2.objects.get(rollno=rollno)
                    setattr(f,code,grade)
                    f.save()
                if(semester=='3'):
                    f =models.ITS3.objects.get(rollno=rollno)
                    setattr(f,code,grade)
                    f.save()
                if(semester=='4'):
                    f =models.ITS4.objects.get(rollno=rollno)
                    setattr(f,code,grade)
                    f.save()
            else:
                return HttpResponse(status=400)
            return HttpResponse(status=200)
        except ObjectDoesNotExist as e:
            return HttpResponse(status=400)

